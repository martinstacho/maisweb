#!/usr/bin/env python3
"""
Generátor testovacích prípadov pre maisweb.
Generuje test-cases.xlsx a test-cases.md zo spoločného zdroja dát.
Spustenie z root repozitára: python3 docs/testing/generate_test_cases.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles.differential import DifferentialStyle
from datetime import datetime
import os

# ─── Brand farby ─────────────────────────────────────────────────────────────
MAIS_ORANGE = 'E66700'
MAIS_DARK   = '1A1A1A'
LIGHT_GRAY  = 'F8F8F8'
WHITE       = 'FFFFFF'

PRIORITY_COLORS = {
    'Critical': 'FFD7D7',
    'High':     'FFE5CC',
    'Medium':   'FFFACD',
    'Low':      'E8F5E9',
}

PRIORITY_FONT_COLORS = {
    'Critical': '8B0000',
    'High':     '7A3800',
    'Medium':   '7A6800',
    'Low':      '1B5E20',
}

_T = Side(style='thin', color='D0D0D0')
BORDER = Border(left=_T, right=_T, top=_T, bottom=_T)

COLUMNS = [
    ('ID',                   12),
    ('Kategória',            15),
    ('Priorita',             12),
    ('Názov',                36),
    ('Predpoklad',           28),
    ('Postup',               56),
    ('Očakávaný výsledok',   48),
    ('Skutočný výsledok',    35),
    ('Stav',                 12),
    ('Poznámky',             30),
    ('Tester',               15),
    ('Dátum',                14),
]

# ─── Test cases ───────────────────────────────────────────────────────────────
# Štruktúra: (ID, Kategória, Priorita, Názov, Predpoklad, Postup, Očakávaný výsledok)
TEST_CASES = [

    # ══════════════════════════════════════════════════════════════════════════
    # FUNKČNÉ (TC-F)
    # ══════════════════════════════════════════════════════════════════════════
    (
        'TC-F-001', 'Funkčné', 'Critical',
        'Otvorenie hlavnej stránky',
        'Internet pripojenie, Chrome/Firefox/Safari',
        '1. Otvor https://www.mais.sk\n2. Počkaj na načítanie (max 5 s)\n3. Otvor DevTools (F12) → Console tab\n4. Over že nie sú červené JS chyby',
        'Stránka sa načíta. Hero blok s nadpisom je viditeľný. Navbar prítomný. Marquee pás sa zobrazuje. Žiadne JS chyby.',
    ),
    (
        'TC-F-002', 'Funkčné', 'High',
        'Scroll linky v navbare',
        'Hlavná stránka /sk je otvorená',
        '1. Klikni "Funkcie" v navbare → over skrolovanie na sekciu Funkcie\n2. Klikni "Školy" → over skrolovanie na sekciu Školy\n3. Klikni "Kontakt" → over skrolovanie na kontaktný blok',
        'Po každom kliknutí sa stránka plynulo skroluje na správnu sekciu. Obsah sekcie je viditeľný.',
    ),
    (
        'TC-F-003', 'Funkčné', 'Critical',
        'Prepínanie jazyka na EN',
        'Aktuálna URL je /sk alebo /sk/...',
        '1. Klikni "EN" v jazykovom prepínači (pravý horný roh)\n2. Over URL adresu v prehliadači\n3. Prečítaj hero nadpis\n4. Over texty v navbare',
        'URL obsahuje /en. Hero nadpis je v angličtine. Navbar položky v EN. Zelená bodka je pri EN.',
    ),
    (
        'TC-F-004', 'Funkčné', 'High',
        'Prepínanie jazyka na UK (ukrajinskú)',
        'Aktuálna URL je /sk alebo /sk/...',
        '1. Klikni "UK" v jazykovom prepínači\n2. Over URL adresu\n3. Prečítaj hero nadpis\n4. Over že text obsahuje ukrajinskú kyrilisku',
        'URL obsahuje /uk. Hero nadpis je v ukrajinskej kyrililike. Všetky texty v ukrajinskej.',
    ),
    (
        'TC-F-005', 'Funkčné', 'High',
        'Prepínanie jazyka na HU (maďarskú)',
        'Aktuálna URL je /sk alebo /sk/...',
        '1. Klikni "HU" v jazykovom prepínači\n2. Over URL adresu\n3. Prečítaj hero nadpis\n4. Skontroluj texty tlačidiel',
        'URL obsahuje /hu. Hero nadpis je v maďarčine. Navbar položky v HU.',
    ),
    (
        'TC-F-006', 'Funkčné', 'Critical',
        'CTA tlačidlo „Nájdi svoju školu"',
        'Hlavná stránka /sk je otvorená, hero blok viditeľný',
        '1. Nájdi primárne (oranžové) CTA tlačidlo v hero\n2. Prečítaj text tlačidla\n3. Klikni naň\n4. Over URL',
        'Tlačidlo má text "Nájdi svoju školu" (alebo ekvivalent). Klik presmeruje na /sk/podpora.',
    ),
    (
        'TC-F-007', 'Funkčné', 'High',
        'CTA tlačidlo „Pre inštitúcie"',
        'Hlavná stránka /sk je otvorená, hero blok viditeľný',
        '1. Nájdi sekundárne CTA tlačidlo v hero\n2. Prečítaj text\n3. Klikni naň\n4. Over URL',
        'Tlačidlo má text "Pre inštitúcie" (alebo ekvivalent). Klik presmeruje na /sk/pre-institucie.',
    ),
    (
        'TC-F-008', 'Funkčné', 'High',
        'Odkaz na web partnerskej školy',
        'Sekcia Školy je viditeľná alebo stránka /podpora',
        '1. Nájdi kartu školy TUKE\n2. Klikni tlačidlo "Webstránka"\n3. Over URL cieľovej stránky\n4. Over že sa otvorí v NOVOM tabe',
        'Otvorí sa web tuke.sk (alebo ekvivalent) v novom tabe. Správna URL pre školu.',
    ),
    (
        'TC-F-009', 'Funkčné', 'High',
        'Odkaz na prihlásenie do MAISu',
        'Sekcia Školy alebo stránka /podpora je viditeľná',
        '1. Nájdi kartu školy TUKE\n2. Klikni tlačidlo "Prihlásenie" alebo "Neviem sa prihlásiť"\n3. Over cieľovú URL',
        'Link vedie na prihlasovaciu stránku MAISu pre danú školu alebo na návod. Otvára sa v novom tabe.',
    ),
    (
        'TC-F-010', 'Funkčné', 'High',
        'Odkaz na e-prihlášku',
        'Sekcia Školy alebo /podpora – karta školy s e-prihláškou (napr. TUKE)',
        '1. Nájdi kartu školy s tlačidlom "E-prihláška"\n2. Klikni naň\n3. Over cieľovú URL\n4. Over nový tab',
        'Link vedie na online prihlášku školy v novom tabe. Školy bez e-prihlášky toto tlačidlo nemajú.',
    ),
    (
        'TC-F-011', 'Funkčné', 'High',
        'Navigácia na stránku /podpora',
        'Aktuálna je hlavná stránka /sk',
        '1. Klikni "Podpora" v navbare\n2. Over URL\n3. Skontroluj obsah',
        'URL sa zmení na /sk/podpora. Stránka zobrazuje karty škôl s tlačidlami. Zobrazuje počet inštitúcií.',
    ),
    (
        'TC-F-012', 'Funkčné', 'High',
        'Navigácia na stránku /pre-institucie',
        'Aktuálna je hlavná stránka /sk',
        '1. Klikni "Pre inštitúcie" v navbare\n2. Over URL\n3. Skontroluj obsah',
        'URL sa zmení na /sk/pre-institucie. Stránka zobrazuje hero, tri dôvody, bento grid a referencie.',
    ),
    (
        'TC-F-013', 'Funkčné', 'Medium',
        'Navigácia na stránku /kontakt',
        'Aktuálna je ľubovoľná stránka',
        '1. Klikni "Kontakt" v navbare alebo footeri\n2. Over URL\n3. Skontroluj obsah',
        'URL sa zmení na /sk/kontakt. Stránka zobrazuje email, telefón a adresu.',
    ),
    (
        'TC-F-014', 'Funkčné', 'Medium',
        'Logo MAIS naviguje na homepage',
        'Aktuálna je podstránka (napr. /sk/podpora)',
        '1. Klikni na logo MAIS v ľavom rohu navbaru\n2. Over URL',
        'Presmeruje na hlavnú stránku /sk. Funguje z ľubovoľnej podstránky.',
    ),
    (
        'TC-F-015', 'Funkčné', 'Medium',
        'Footer navigačné linky',
        'Footer je viditeľný (najspodnejšia časť stránky)',
        '1. Nájdi footer\n2. Klikni na každý navigačný link vo footeri\n3. Over ciele',
        'Každý footer link vedie na správnu stránku. Externé linky v novom tabe.',
    ),
    (
        'TC-F-016', 'Funkčné', 'Medium',
        'Emailový link mais@mais.sk',
        'Zariadenie má nastavený emailový klient',
        '1. Nájdi link mais@mais.sk (footer alebo /kontakt)\n2. Klikni naň',
        'Otvorí sa emailový klient s predvyplnenou adresou mais@mais.sk. Správna adresa.',
    ),
    (
        'TC-F-017', 'Funkčné', 'Medium',
        'Telefónny link +421 915 724 757',
        'Mobilné zariadenie alebo desktop s VoIP',
        '1. Nájdi telefónne číslo na stránke\n2. Klikni naň',
        'Otvorí sa dialóg pre volanie s číslom +421915724757. Na desktop VoIP aplikácia.',
    ),
    (
        'TC-F-018', 'Funkčné', 'High',
        '/pre-institucie CTA „Kontaktujte nás"',
        'Stránka /sk/pre-institucie je otvorená',
        '1. Skroluj na CTA sekciu na spodku stránky\n2. Klikni tlačidlo "Kontaktujte nás"\n3. Over kam vedie',
        'Presmeruje na /sk/kontakt alebo otvorí emailový klient. Kontaktné údaje dostupné.',
    ),
    (
        'TC-F-019', 'Funkčné', 'High',
        '/podpora – kontaktný banner',
        'Stránka /sk/podpora je otvorená',
        '1. Skroluj na spodok stránky\n2. Nájdi banner "Tvoja škola nie je v zozname?"\n3. Klikni na kontaktné tlačidlo',
        'Banner je viditeľný. Klik otvorí emailový klient alebo vedie na /kontakt.',
    ),
    (
        'TC-F-020', 'Funkčné', 'High',
        'Hamburger menu na mobile',
        'Prehliadač v mobilnom zobrazení (≤ 768 px) alebo skutočný mobil',
        '1. Zmenši okno na 375 px (F12 → DevTools → Toggle device)\n2. Klikni hamburger ikonu (≡) v navbare\n3. Over zobrazenie menu\n4. Klikni na položku\n5. Over zatváranie',
        'Menu sa otvorí s viditeľnými položkami. Klik naviguje správne. Menu sa zavrie po výbere.',
    ),

    # ══════════════════════════════════════════════════════════════════════════
    # VIZUÁLNE (TC-V)
    # ══════════════════════════════════════════════════════════════════════════
    (
        'TC-V-001', 'Vizuálne', 'High',
        'Desktop layout 1920×1080',
        'Prehliadač nastavený na 1920 × 1080 px',
        '1. Otvor https://www.mais.sk na 1920 px\n2. Skroluj cez celú stránku\n3. Skontroluj: hero, marquee, stats, features, orbital, integrácie, školy, footer',
        'Hero na celú šírku. Marquee viditeľný. Stats v 4 stĺpcoch. Features 3×2. Nič nepretéka.',
    ),
    (
        'TC-V-002', 'Vizuálne', 'Medium',
        'Notebook layout 1366×768',
        'Prehliadač nastavený na 1366 × 768 px',
        '1. Otvor stránku na 1366 px\n2. Skroluj cez celú stránku\n3. Over že nič nie je orezané alebo pretekajúce',
        'Všetky sekcie viditeľné. Texty čitateľné. Žiadny horizontálny scrollbar.',
    ),
    (
        'TC-V-003', 'Vizuálne', 'High',
        'Tablet layout 768 px',
        'DevTools otvorené, šírka nastavená na 768 px',
        '1. F12 → Toggle device toolbar\n2. Nastav šírku 768 px\n3. Skontroluj navbar\n4. Over zalomenie kariet\n5. Over čitateľnosť textov',
        'Navbar skrytý do hamburger menu. Karty partnrov v 2 stĺpcoch. Texty čitateľné.',
    ),
    (
        'TC-V-004', 'Vizuálne', 'Critical',
        'Mobil layout 375 px (iPhone SE)',
        'DevTools – iPhone SE (375 × 667)',
        '1. F12 → Toggle device → iPhone SE\n2. Skroluj cez celú stránku\n3. Over hamburger menu\n4. Over veľkosť tlačidiel (min 44 × 44 px)\n5. Over overflow',
        'Všetko v 1 stĺpci. Žiadny horizontálny scroll. Tlačidlá dotykovo prístupné. Texty čitateľné.',
    ),
    (
        'TC-V-005', 'Vizuálne', 'Medium',
        'Hero gradient / background animácia',
        'Desktop, stránka sa práve načítala',
        '1. Otvor hlavnú stránku\n2. Pozoruj hero blok 5 sekúnd\n3. Over pohyb alebo zmenu gradientu pozadia',
        'Gradient v hero sa animuje. Animácia plynulá, bez zásekov.',
    ),
    (
        'TC-V-006', 'Vizuálne', 'Medium',
        'Marquee pás – kontinuálna animácia',
        'Hlavná stránka načítaná',
        '1. Pozoruj marquee pás (pod hero)\n2. Sleduj animáciu 5 sekúnd\n3. Over smer a plynulosť\n4. Over že sa opakuje v slučke',
        'Marquee sa kontinuálne posúva. Školy sa opakujú v slučke bez prerušení.',
    ),
    (
        'TC-V-007', 'Vizuálne', 'Medium',
        'NumberTicker – animácia čísel v stats',
        'Stats sekcia je v zornom poli prehliadača',
        '1. Skroluj na stats sekciu\n2. Sleduj čísla 22+, 9, 50 000+, 100%\n3. Over animáciu počítania od 0',
        'Čísla sa animujú od 0 pri prvom zobrazení v zornom poli. Zastavujú na správnych hodnotách.',
    ),
    (
        'TC-V-008', 'Vizuálne', 'Low',
        'Orbital animácia – moduly obiehajú',
        'Sekcia Architektúra je viditeľná (desktop)',
        '1. Skroluj na sekciu Architektúra\n2. Pozoruj orbital animáciu\n3. Over rotáciu modulov\n4. Over text v strede',
        '9 modulov obiehajú okolo "MAIS core engine" v strede. Animácia plynulá. Text čitateľný.',
    ),
    (
        'TC-V-009', 'Vizuálne', 'Medium',
        'Hover stav tlačidiel',
        'Desktop s myšou, tlačidlá viditeľné',
        '1. Najeď myšou na primárne oranžové tlačidlo\n2. Over vizuálnu zmenu (farba, tieň)\n3. Odjeď a over návrat\n4. Opakuj pre sekundárne tlačidlo',
        'Tlačidlá majú viditeľný hover efekt. Stav sa vráti po odchode kurzora.',
    ),
    (
        'TC-V-010', 'Vizuálne', 'Medium',
        'Focus stav – klávesnicová navigácia',
        'Prehliadač s klávesnicou',
        '1. Klikni na prázdne miesto na stránke\n2. Stlač Tab\n3. Sleduj focus indikátor (modrý/biely outline)\n4. Prejdi Tab cez všetky interaktívne prvky',
        'Každý interaktívny element má viditeľný focus ring. Poradie navigácie je logické.',
    ),
    (
        'TC-V-011', 'Vizuálne', 'Low',
        'Aktívna bodka v language switcheri',
        'Stránka /sk je otvorená',
        '1. Pozri language switcher v navbare\n2. Over ktorý jazyk má farebnú bodku\n3. Prepni na EN\n4. Over polohu bodky',
        'Aktívny jazyk (SK) má zelenú bodku. Po prepnutí na EN sa bodka presunie na EN.',
    ),
    (
        'TC-V-012', 'Vizuálne', 'Low',
        'Navbar efekt pri skrolovaní',
        'Hlavná stránka, hero je viditeľný (navrchu)',
        '1. Pozoruj navbar pri načítaní\n2. Začni skrolovať nadol\n3. Over zmenu pozadia / tieňa navbaru\n4. Over sticky polohu',
        'Navbar sa pri skrolovaní zmení (blur/pozadie). Zostáva sticky na vrchu stránky.',
    ),
    (
        'TC-V-013', 'Vizuálne', 'Low',
        'Hover efekt na kartách partnerov',
        'Desktop s myšou, sekcia Školy viditeľná',
        '1. Najeď myšou na kartu školy\n2. Over hover efekt (tieň, scale, border)\n3. Odjeď a over návrat',
        'Karta má vizuálny hover efekt (tieň alebo transform). Animácia plynulá.',
    ),
    (
        'TC-V-014', 'Vizuálne', 'Medium',
        'AOS scroll animácie',
        'Prehliadač bez prefers-reduced-motion',
        '1. Skroluj pomaly cez celú stránku od hora nadol\n2. Sleduj vstup sekcií do zorného poľa',
        'Sekcie vstupujú s animáciou (fade-up, fade-in). Animácie plynulé, bez zásekov.',
    ),
    (
        'TC-V-015', 'Vizuálne', 'Medium',
        'Integrácie chipy – farby podľa kategórie',
        'Sekcia Integrácie viditeľná na hlavnej stránke',
        '1. Nájdi sekciu Integrácie\n2. Over 6 kategórií (Identita, Ekonomika, Štúdium, Registratúra, Externé, Mobilné)\n3. Over rôzne farby chipov pre rôzne kategórie',
        'Každá kategória má odlišnú farebnú schému. Chipy čitateľné. Farby konzistentné v rámci kategórie.',
    ),

    # ══════════════════════════════════════════════════════════════════════════
    # OBSAHOVÉ (TC-O)
    # ══════════════════════════════════════════════════════════════════════════
    (
        'TC-O-001', 'Obsahové', 'Critical',
        'Hero text SK',
        'Stránka /sk načítaná',
        '1. Prečítaj hlavný nadpis hero sekcie\n2. Prečítaj podnadpis\n3. Skontroluj badge nad nadpisom',
        'Nadpis: "Akademický informačný systém, ktorý poháňa slovenské univerzity" (alebo aktuálny z DB). Podnadpis v SK. Badge viditeľný.',
    ),
    (
        'TC-O-002', 'Obsahové', 'High',
        'Hero text EN',
        'Stránka /en načítaná',
        '1. Prečítaj hlavný nadpis hero sekcie\n2. Prečítaj podnadpis\n3. Over že nie sú slovenské slová',
        'Nadpis v angličtine. Podnadpis v angličtine. Žiadne slovenské slová v hero.',
    ),
    (
        'TC-O-003', 'Obsahové', 'High',
        'Hero text UK (kyrilika)',
        'Stránka /uk načítaná',
        '1. Prečítaj hlavný nadpis hero sekcie\n2. Over že text je kyrilika\n3. Over čitateľnosť fontu',
        'Nadpis v ukrajinskej kyrililike. Font zobrazuje cyrilické znaky správne.',
    ),
    (
        'TC-O-004', 'Obsahové', 'High',
        'Hero text HU (maďarčina)',
        'Stránka /hu načítaná',
        '1. Prečítaj hlavný nadpis hero sekcie\n2. Over maďarské diakritické znaky (é, á, ő, ű)',
        'Nadpis v maďarčine. Diakritické znaky sa zobrazujú správne.',
    ),
    (
        'TC-O-005', 'Obsahové', 'Critical',
        'Stats sekcia – 4 správne hodnoty',
        'Stats sekcia viditeľná na hlavnej stránke',
        '1. Nájdi sekciu so 4 veľkými číslami\n2. Prečítaj hodnotu 1 (roky)\n3. Prečítaj hodnotu 2 (inštitúcie)\n4. Prečítaj hodnotu 3 (používatelia)\n5. Prečítaj hodnotu 4 (kompatibilita)',
        '"22+" rokov skúseností. "9" inštitúcií. "50 000+" aktívnych používateľov. "100%" kompatibilita s legislatívou.',
    ),
    (
        'TC-O-006', 'Obsahové', 'High',
        'Marquee pás – 9 partnerských škôl',
        'Marquee pás viditeľný pod hero',
        '1. Pozoruj marquee pás\n2. Vypíš všetky skratky škôl\n3. Over počet (9)',
        'Všetkých 9 škôl prítomných: TUKE, TRUNI, SZU, UNIPO, AOS, APZ, BISLA, DTI, VŠBM. Každá má bodku a mesto.',
    ),
    (
        'TC-O-007', 'Obsahové', 'Medium',
        'Integrácie – počet položiek a kategórie',
        'Sekcia Integrácie viditeľná na hlavnej stránke',
        '1. Spočítaj integračné chipy\n2. Over 6 kategórií\n3. Nájdi LDAP, SAP, CRŠ ako konkrétne príklady',
        '29 integrácií v 6 kategóriách. Identita (LDAP, OAuth2, SSO), Ekonomika (SAP, Štátna pokladnica), Štúdium (CRŠ, CVTI), Registratúra, Externé, Mobilné.',
    ),
    (
        'TC-O-008', 'Obsahové', 'Medium',
        'Partneri – mestá a logá/monogramy',
        'Sekcia Školy alebo /skoly viditeľná',
        '1. Pre každú kartu školy over: logo/monogram, plný názov, mesto\n2. Over správnosť miest: TUKE = Košice, TRUNI = Trnava, SZU = Bratislava',
        'Všetkých 9 škôl má logo/monogram, plný názov a správne mesto. Logá sa načítavajú.',
    ),
    (
        'TC-O-009', 'Obsahové', 'Critical',
        'Kontaktné info (email a telefón)',
        'Footer alebo /kontakt viditeľný',
        '1. Nájdi emailovú adresu\n2. Prečítaj presný text\n3. Nájdi telefónne číslo\n4. Prečítaj presný text',
        'Email: mais@mais.sk. Telefón: +421 915 724 757. Oba sú klikateľné linky.',
    ),
    (
        'TC-O-010', 'Obsahové', 'Low',
        'Copyright text v footeri',
        'Footer viditeľný',
        '1. Nájdi copyright text v footeri\n2. Prečítaj rok a názov spoločnosti',
        'Footer obsahuje copyright s aktuálnym rokom (2026) a názvom ITernal s.r.o.',
    ),
    (
        'TC-O-011', 'Obsahové', 'High',
        'Feature karty – 6 modulov MAIS',
        'Sekcia Funkcie/Features viditeľná',
        '1. Spočítaj feature karty\n2. Prečítaj nadpis každej karty\n3. Over že každá má popis',
        '6 kariet: Študijná agenda, E-prihláška, Bezpečnosť, Výkon, Integrácie, Podpora. Každá má ikonu, nadpis a popis.',
    ),
    (
        'TC-O-012', 'Obsahové', 'High',
        '/podpora – texty a tlačidlá škôl',
        'Stránka /sk/podpora načítaná',
        '1. Prečítaj hero text na /podpora\n2. Over live číslo inštitúcií (malo by byť 9)\n3. Skontroluj texty tlačidiel na kartách škôl',
        'Hero text v SK. Live číslo = 9. Tlačidlá: "Webstránka", "Prihlásenie", "E-prihláška".',
    ),
    (
        'TC-O-013', 'Obsahové', 'High',
        '/pre-institucie – referencie (3 citáty)',
        'Stránka /sk/pre-institucie načítaná',
        '1. Skroluj na sekciu Referencie\n2. Spočítaj citáty\n3. Prečítaj každý citát\n4. Over meno, titul a školu',
        '3 citáty od zástupcov škôl. Každý má meno, titul a školu. Text zmysluplný v SK.',
    ),
    (
        'TC-O-014', 'Obsahové', 'Medium',
        '/kontakt – kompletné údaje',
        'Stránka /sk/kontakt načítaná',
        '1. Prečítaj email\n2. Prečítaj telefón\n3. Prečítaj adresu\n4. Over spoločnosť',
        'Email: mais@mais.sk. Telefón: +421 915 724 757. Adresa: Dubnica nad Váhom. Spoločnosť: ITernal s.r.o.',
    ),
    (
        'TC-O-015', 'Obsahové', 'Medium',
        'Orbital animácia – 9 modulov MAIS',
        'Sekcia Architektúra viditeľná (desktop)',
        '1. Nájdi orbital animáciu\n2. Spočítaj moduly\n3. Prečítaj text v strede\n4. Over 4 technické detaily pod animáciou',
        '9 modulov okolo "MAIS core engine". V strede text "MAIS core engine". Pod animáciou 4 detaily: SSO, databáza, integrácie, zálohy.',
    ),

    # ══════════════════════════════════════════════════════════════════════════
    # ADMIN (TC-A)
    # ══════════════════════════════════════════════════════════════════════════
    (
        'TC-A-001', 'Admin', 'Critical',
        'Login so správnymi credentials',
        'Poznáš prihlasovacie údaje',
        '1. Otvor https://www.mais.sk/admin/login\n2. Zadaj email: admin@mais.sk\n3. Zadaj správne heslo\n4. Klikni "Prihlásiť sa"',
        'Presmeruje na /admin dashboard. Admin rozhranie sa zobrazí s navigačnými sekciami.',
    ),
    (
        'TC-A-002', 'Admin', 'Critical',
        'Login s nesprávnym heslom',
        'Admin login stránka otvorená',
        '1. Zadaj email: admin@mais.sk\n2. Zadaj nesprávne heslo (napr. wrongpass99!)\n3. Klikni "Prihlásiť sa"',
        'Login zlyhá. Zobrazí sa chybová správa. URL ostáva /admin/login. Formulár ostáva viditeľný.',
    ),
    (
        'TC-A-003', 'Admin', 'High',
        'Rate limiting – blokovanie po 6 pokusoch',
        'Admin login stránka otvorená (testovať na beta)',
        '1. Zadaj nesprávne heslo 5× v rade\n2. Skontroluj správu po 5. pokuse\n3. Zadaj nesprávne heslo 6×\n4. Skontroluj správu po 6. pokuse',
        'Po 5 pokusoch: štandardná chyba. Po 6. pokuse: "Príliš veľa pokusov prihlásenia, skúste neskôr."',
    ),
    (
        'TC-A-004', 'Admin', 'Critical',
        'Force password change pri prvom prihlásení',
        'Nový admin účet s mustChangePassword=true, alebo default heslo',
        '1. Prihlás sa s účtom mustChangePassword=true\n2. Sleduj presmerovanie\n3. Pokús sa otvoriť /admin priamo bez zmeny hesla',
        'Automaticky presmeruje na /admin/change-password. Priamy prístup na /admin nie je možný bez zmeny hesla.',
    ),
    (
        'TC-A-005', 'Admin', 'High',
        'Validácia hesla – príliš krátke',
        'Stránka /admin/change-password otvorená',
        '1. Zadaj aktuálne heslo (správne)\n2. Nové heslo: "abc"\n3. Potvrď: "abc"\n4. Klikni Uložiť',
        'Chyba: "Heslo musí mať aspoň 8 znakov". Heslo sa neuloží. Zostávaš na stránke.',
    ),
    (
        'TC-A-006', 'Admin', 'High',
        'Validácia hesla – chýbajúce typy znakov',
        'Stránka /admin/change-password otvorená',
        '1. Zadaj aktuálne heslo (správne)\n2. Nové heslo: "heslo123" (chýba veľké písmeno a špeciálny znak)\n3. Klikni Uložiť',
        'Chyby: "Musí obsahovať veľké písmeno" + "Musí obsahovať špeciálny znak". Heslo neuložené.',
    ),
    (
        'TC-A-007', 'Admin', 'Critical',
        'Pridanie nového partnera',
        'Prihlásený do adminu, /admin/partners otvorená',
        '1. Klikni "Pridať partnera"\n2. Vyplň: skratka, plný názov, mesto, URL webu, URL prihlásenia\n3. Klikni Uložiť\n4. Over v zozname\n5. Otvor web a over kartu školy',
        'Partner sa uloží. V zozname partnerov viditeľný. Na webe (/sk alebo /podpora) sa zobrazí nová karta školy.',
    ),
    (
        'TC-A-008', 'Admin', 'High',
        'Editácia existujúceho partnera',
        'Prihlásený do adminu, existuje aspoň 1 partner',
        '1. V zozname klikni Upraviť na partnerovi\n2. Zmeň mestý (pridaj "TEST")\n3. Ulož\n4. Over zmenu v zozname\n5. Otvor web a over zmenu',
        'Zmena sa uloží. Admin zoznam zobrazuje novú hodnotu. Web okamžite zobrazuje zmenu.',
    ),
    (
        'TC-A-009', 'Admin', 'High',
        'Zmazanie partnera',
        'Prihlásený do adminu, existuje testovací partner',
        '1. V zozname klikni Zmazať na testovacím partnerovi\n2. Over confirmation dialóg\n3. Potvrď zmazanie\n4. Over zoznam\n5. Otvor web a over',
        'Confirmation dialóg sa zobrazí. Po potvrdení partner zmizne zo zoznamu. Web kartu školy viac nezobrazuje.',
    ),
    (
        'TC-A-010', 'Admin', 'High',
        'CRUD operácie pre integrácie',
        'Prihlásený do adminu',
        '1. Otvor /admin/integrations\n2. Pridaj integráciu (názov "TEST-INTEG", kategória)\n3. Over v zozname\n4. Uprav ju\n5. Zmaž ju\n6. Over web',
        'Integrácia sa pridá, upraví a zmaže. Web v sekcii Integrácie reflektuje každú zmenu.',
    ),
    (
        'TC-A-011', 'Admin', 'Medium',
        'CRUD operácie pre referencie',
        'Prihlásený do adminu',
        '1. Otvor /admin/testimonials\n2. Pridaj referenciu (citát, meno, titul, škola)\n3. Over v zozname\n4. Uprav ju\n5. Zmaž ju\n6. Over /pre-institucie',
        'Referencia sa pridá, upraví a zmaže. Sekcia Referencie na /pre-institucie reflektuje zmeny.',
    ),
    (
        'TC-A-012', 'Admin', 'High',
        'Root admin pridáva nového správcu',
        'Prihlásený ako root admin (sekcia Správcovia viditeľná)',
        '1. Otvor /admin/users\n2. Klikni Pridať správcu\n3. Zadaj email nového správcu a heslo\n4. Ulož\n5. Over v zozname',
        'Nový správca vytvorený. Viditeľný v zozname používateľov s rolou "správca".',
    ),
    (
        'TC-A-013', 'Admin', 'High',
        'Non-root správca nevidí sekciu Správcovia',
        'Prihlásený ako non-root správca',
        '1. Prihlás sa ako non-root správca\n2. Pozri sidebar admin rozhrania\n3. Hľadaj sekciu "Správcovia" alebo "Používatelia"',
        'Sekcia Správcovia NIE JE viditeľná v sidebar. Non-root nemá prístup k správe účtov.',
    ),
    (
        'TC-A-014', 'Admin', 'Medium',
        'Audit log zaznamenáva admin akcie',
        'Prihlásený ako root admin',
        '1. Vykonaj akciu: uprav partnera\n2. Otvor /admin/audit\n3. Nájdi záznam akcie\n4. Over: akcia, email, čas',
        'Audit log zobrazuje záznam. Správna akcia, email administrátora a čas. Zoradené od najnovšieho.',
    ),
    (
        'TC-A-015', 'Admin', 'Medium',
        'Session expirácia po dlhej inaktivite',
        'Prihlásený do adminu (admin session trvá 4 h)',
        '1. Prihláš sa do adminu\n2. Počkaj 4+ hodiny bez aktivity (alebo manuálne expiruj cookie v DevTools)\n3. Pokús sa vykonať akciu alebo načítaj /admin',
        'Session je expirovaná. Presmeruje na /admin/login. Chráneného obsahu nie je možné dostať sa bez opätovného prihlásenia.',
    ),

    # ══════════════════════════════════════════════════════════════════════════
    # CONTENT CMS (TC-C)
    # ══════════════════════════════════════════════════════════════════════════
    (
        'TC-C-001', 'Content CMS', 'High',
        'Otvorenie /admin/sections – prehľad sekcií',
        'Prihlásený do adminu',
        '1. Klikni "Sekcie" alebo "Obsah webu" v admin navbare\n2. Over počet dlaždíc\n3. Over popis každej dlaždice',
        '4 dlaždice: Homepage, Pre inštitúcie, Podpora, Footer & Kontakt. Každá ukazuje počet skupín a polí.',
    ),
    (
        'TC-C-002', 'Content CMS', 'Critical',
        'Editácia hero.title1 SK',
        'Prihlásený, /admin/sections/homepage otvorená, záložka SK',
        '1. Nájdi skupinu "Hero"\n2. Klikni na pole "hero.title1"\n3. Vymaž obsah a napíš "TEST NADPIS"\n4. Sleduj status indikátor\n5. Klikni "Uložiť skupinu"',
        'Po zmene: status "neuložené". Po uložení: status "upravené v DB". Toast "Uložené ✓" zelený.',
    ),
    (
        'TC-C-003', 'Content CMS', 'Critical',
        'CMS zmena sa prejaví na webe',
        'TC-C-002 dokončený – hero.title1 SK zmenený na "TEST NADPIS"',
        '1. Otvor novú záložku prehliadača\n2. Otvor https://www.mais.sk/sk\n3. Prečítaj hero nadpis',
        'Hero nadpis zobrazuje "TEST NADPIS". Nie pôvodnú hodnotu z messages.json.',
    ),
    (
        'TC-C-004', 'Content CMS', 'High',
        'Reset textu na pôvodnú hodnotu',
        'TC-C-002 dokončený – pole hero.title1 SK je zmenené',
        '1. Nájdi zmenené pole v admin\n2. Klikni "Reset" pri danom poli\n3. Over status poľa\n4. Ulož skupinu\n5. Otvor web a over nadpis',
        'Po resete: status "pôvodné". Text = pôvodná hodnota z messages.json. Web zobrazuje pôvodnú hodnotu.',
    ),
    (
        'TC-C-005', 'Content CMS', 'High',
        'Editácia textu v iných jazykoch (EN)',
        'Prihlásený, /admin/sections/homepage otvorená',
        '1. Prepni záložku na "EN"\n2. Nájdi pole hero.title1\n3. Zmeň na "TEST EN HEADING"\n4. Ulož\n5. Otvor /en na webe\n6. Otvor /sk a over že SK sa nezmenilo',
        'Zmena viditeľná len na /en. SK, UK, HU zostávajú nezmenené.',
    ),
    (
        'TC-C-006', 'Content CMS', 'Medium',
        'Status indikátory – hierarchia',
        'Prihlásený, sekcia homepage otvorená',
        '1. Over status poľa bez zmeny (= "pôvodné")\n2. Zmeň hodnotu → status = "neuložené"\n3. Ulož → status = "upravené v DB"\n4. Over status skupiny a dlaždice stránky',
        'Status propaguje hierarchicky: pole → skupina → stránka. Každý level odráža reálny stav.',
    ),
    (
        'TC-C-007', 'Content CMS', 'Medium',
        'Multi-field save (viacero polí naraz)',
        'Prihlásený, sekcia homepage otvorená',
        '1. Zmeň 3 rôzne polia v jednej skupine\n2. Klikni raz "Uložiť skupinu"\n3. Over toast\n4. Over status všetkých 3 polí',
        'Všetky 3 polia uložené jednou akciou. Toast "Uložené ✓". Status každého = "upravené v DB".',
    ),
    (
        'TC-C-008', 'Content CMS', 'Low',
        'Dlhý text nepokazí layout webu',
        'Prihlásený do adminu',
        '1. Do CMS poľa zadaj text dlhší ako 300 znakov\n2. Ulož\n3. Otvor web na desktop (1366 px) a mobile (375 px)\n4. Over layout',
        'Layout sa nepokazí. Text je zabalený alebo skrátený. Žiadny overflow mimo kontajnera.',
    ),
    (
        'TC-C-009', 'Content CMS', 'Medium',
        'Špeciálne znaky v CMS texte',
        'Prihlásený do adminu',
        '1. Do CMS poľa zadaj: & < > " \' a emoji 🎓\n2. Ulož\n3. Otvor web a over zobrazenie\n4. Over DevTools Console – žiadna JS chyba',
        'Špeciálne znaky zobrazené správne (& ako &, < ako <). Emoji viditeľné. Žiadna JS chyba.',
    ),
    (
        'TC-C-010', 'Content CMS', 'Low',
        'Toast notifikácie pri ukladaní a resetovaní',
        'Prihlásený, CMS zmena pripravená na uloženie',
        '1. Klikni "Uložiť skupinu"\n2. Over toast správu a farbu\n3. Počkaj kým zmizne\n4. Klikni "Reset" na poli\n5. Over toast',
        'Uloženie: zelený toast "Uložené ✓", automaticky zmizne. Reset: toast s informáciou. Žiadne toast prekrývanie.',
    ),

    # ══════════════════════════════════════════════════════════════════════════
    # SEO (TC-S)
    # ══════════════════════════════════════════════════════════════════════════
    (
        'TC-S-001', 'SEO', 'High',
        'robots.txt – obsah a pravidlá',
        'Web dostupný',
        '1. Otvor https://www.mais.sk/robots.txt\n2. Over Allow pravidlá\n3. Over Disallow pravidlá\n4. Over sitemap link\n5. Over GPTBot pravidlo',
        'Allow: /sk, /en, /uk, /hu, /skoly. Disallow: /admin, /api. Sitemap link prítomný. GPTBot Disallow: /.',
    ),
    (
        'TC-S-002', 'SEO', 'High',
        'sitemap.xml – 12 URL so hreflang',
        'Web dostupný',
        '1. Otvor https://www.mais.sk/sitemap.xml\n2. Spočítaj <url> elementy\n3. Over hreflang alternates v každom URL\n4. Over lastmod dátum',
        '12 URL (4 jazyky × 3 stránky). Každé URL má 4 hreflang alternáty + x-default. Lastmod prítomný.',
    ),
    (
        'TC-S-003', 'SEO', 'High',
        'Hreflang tagy v <head>',
        'Web dostupný',
        '1. Otvor https://www.mais.sk/sk\n2. Ctrl+U (View source)\n3. Hľadaj: <link rel="alternate" hreflang=\n4. Spočítaj a over hodnoty href',
        '5 hreflang tagov: hreflang="sk", "en", "uk", "hu", "x-default". Každý s korektnou URL.',
    ),
    (
        'TC-S-004', 'SEO', 'High',
        'Meta title a description lokalizované',
        'Web dostupný v 4 jazykoch',
        '1. Otvor /sk → DevTools → Elements → hľadaj <title>\n2. Opakuj pre /hu\n3. Over description tag pre oba jazyky',
        '/sk: title a description v slovenčine. /hu: title a description v maďarčine. Každý jazyk má unikátny obsah.',
    ),
    (
        'TC-S-005', 'SEO', 'Medium',
        'OG image dynamický endpoint',
        'Web dostupný',
        '1. Otvor https://www.mais.sk/sk/opengraph-image\n2. Over Content-Type odpovede\n3. Over rozmery obrázka',
        'Server vracia PNG obrázok. Content-Type: image/png. Rozmer 1200 × 630 px.',
    ),
    (
        'TC-S-006', 'SEO', 'High',
        'JSON-LD structured data',
        'Web dostupný',
        '1. Otvor https://www.mais.sk/sk\n2. Ctrl+U (View source)\n3. Hľadaj <script type="application/ld+json">\n4. Skontroluj @type hodnotu',
        '<script type="application/ld+json"> prítomný. @type: Organization a/alebo SoftwareApplication. name, url, description vyplnené.',
    ),
    (
        'TC-S-007', 'SEO', 'High',
        'Admin stránky – noindex',
        'DevTools dostupné',
        '1. DevTools → Network tab\n2. Načítaj /admin/login\n3. Klikni na HTML dokument požiadavku\n4. Pozri Response Headers',
        'Header X-Robots-Tag: noindex, nofollow, noarchive prítomný v odpovedi /admin stránok.',
    ),
    (
        'TC-S-008', 'SEO', 'Medium',
        'Canonical URL v <head>',
        'Web dostupný',
        '1. Otvor https://www.mais.sk/sk\n2. Ctrl+U (View source)\n3. Hľadaj <link rel="canonical"\n4. Over hodnotu href',
        '<link rel="canonical" href="https://www.mais.sk/sk"> prítomný. URL súhlasí s aktuálnou stránkou.',
    ),

    # ══════════════════════════════════════════════════════════════════════════
    # SECURITY (TC-SEC)
    # ══════════════════════════════════════════════════════════════════════════
    (
        'TC-SEC-001', 'Security', 'High',
        'Security headers na verejných stránkach',
        'DevTools dostupné',
        '1. DevTools → Network tab\n2. Načítaj /sk\n3. Klikni na HTML dokument v Network\n4. Pozri Response Headers\n5. Hľadaj bezpečnostné hlavičky',
        'Prítomné: Strict-Transport-Security, X-Frame-Options: DENY, X-Content-Type-Options: nosniff, Referrer-Policy, Permissions-Policy.',
    ),
    (
        'TC-SEC-002', 'Security', 'High',
        'Admin no-cache hlavičky',
        'Admin dostupný, DevTools otvorené',
        '1. DevTools → Network tab\n2. Načítaj /admin/login\n3. Klikni na HTML dokument\n4. Over Cache-Control header',
        'Cache-Control: no-store, no-cache, must-revalidate prítomný na /admin stránkach.',
    ),
    (
        'TC-SEC-003', 'Security', 'High',
        'Admin X-Robots-Tag noindex v HTTP odpovedi',
        'DevTools dostupné',
        '1. DevTools → Network tab\n2. Načítaj /admin alebo /admin/login\n3. Pozri Response Headers',
        'X-Robots-Tag: noindex, nofollow, noarchive prítomný v HTTP odpovedi /admin ciest.',
    ),
    (
        'TC-SEC-004', 'Security', 'Medium',
        'HTTPS redirect (HTTP → HTTPS)',
        'Produkčný server mais.sk je aktívny',
        '1. Zadaj do prehliadača: http://www.mais.sk\n2. Sleduj presmerovanie\n3. Over finálnu URL',
        'Automaticky presmeruje na https://www.mais.sk. URL sa zmení na HTTPS. Žiadne bezpečnostné varovania.',
    ),
    (
        'TC-SEC-005', 'Security', 'High',
        'Brute force – blokovanie po 6 pokusoch',
        'Admin login dostupný (testovať na beta, nie produkcia)',
        '1. Otvor /admin/login\n2. Zadaj nesprávne heslo 5×\n3. Over správu po 5. pokuse\n4. Zadaj nesprávne heslo 6×\n5. Over správu',
        'Prvých 5 pokusov: štandardná chybová správa. Po 6. pokuse: správa o blokovaní. IP zablokovaná.',
    ),
    (
        'TC-SEC-006', 'Security', 'Medium',
        'SQL injection – odolnosť login formulára',
        'Admin login stránka otvorená',
        '1. Email: admin\' OR \'1\'=\'1\n2. Heslo: anything123!\n3. Odošli formulár\n4. Over odpoveď',
        'Login zlyhá normálne (nesprávne credentials). Žiadna DB chyba ani neočakávaná odpoveď. App ostáva stabilná.',
    ),
    (
        'TC-SEC-007', 'Security', 'High',
        'XSS ochrana v CMS poliach',
        'Prihlásený do adminu, /admin/sections otvorená',
        '1. Do CMS textového poľa zadaj: <script>alert(\'XSS\')</script>\n2. Ulož\n3. Otvor web v novom tabe\n4. Sleduj či sa spustí alert',
        'Alert sa NESPUSTÍ. Text je zobrazený ako čistý text alebo je escapenutý. Žiadne JavaScript execution.',
    ),

    # ══════════════════════════════════════════════════════════════════════════
    # LOKALIZÁCIA (TC-L)
    # ══════════════════════════════════════════════════════════════════════════
    (
        'TC-L-001', 'Lokalizácia', 'High',
        'Auto-detekcia maďarského jazyka prehliadača',
        'Chrome prehliadač, NEXT_LOCALE cookie neexistuje',
        '1. Chrome → Nastavenia → Jazyky → Pridaj Hungarian (hu) ako primárny\n2. Otvor nové inkognito okno\n3. Otvor https://www.mais.sk/\n4. Over URL presmerovanie',
        'Automaticky presmeruje na /hu. Stránka sa zobrazí v maďarčine.',
    ),
    (
        'TC-L-002', 'Lokalizácia', 'Medium',
        'Cookie persistence – zachovanie jazykovej voľby',
        'Web otvorený v prehliadači',
        '1. Klikni SK v prepínači\n2. DevTools → Application → Cookies → mais.sk\n3. Nájdi NEXT_LOCALE cookie\n4. Zavri a znova otvor prehliadač\n5. Otvor mais.sk',
        'NEXT_LOCALE=sk cookie nastavená. Pri ďalšej návšteve sa načíta /sk bez ohľadu na jazyk prehliadača.',
    ),
    (
        'TC-L-003', 'Lokalizácia', 'Medium',
        'Fallback na SK pre nepodporovaný jazyk',
        'Chrome prehliadač',
        '1. Chrome → Nastavenia → Jazyky → Nastav German (de) ako primárny\n2. Vymaž NEXT_LOCALE cookie\n3. Otvor mais.sk v inkognito okne',
        'Presmeruje na /sk (fallback). Stránka sa zobrazí v slovenčine.',
    ),
    (
        'TC-L-004', 'Lokalizácia', 'Medium',
        'Prepínanie jazyka zachováva podstránku',
        'Aktuálna URL je /sk/podpora',
        '1. Otvor /sk/podpora\n2. Klikni EN v prepínači\n3. Over URL',
        'URL sa zmení na /en/podpora. Zostaneš na stránke /podpora. Nie presmerovaný na homepage.',
    ),
    (
        'TC-L-005', 'Lokalizácia', 'High',
        'URL prefix pre všetky 4 jazyky a podstránky',
        'Web dostupný',
        '1. Over /sk, /en, /uk, /hu (hlavná)\n2. Over /sk/podpora, /en/podpora, /uk/podpora, /hu/podpora\n3. Over /sk/pre-institucie pre každý jazyk',
        'Každý jazyk má prefix /sk/, /en/, /uk/, /hu/. Všetky kombinácie jazyk+stránka fungujú a vracajú 200.',
    ),

    # ══════════════════════════════════════════════════════════════════════════
    # EDGE CASES (TC-E)
    # ══════════════════════════════════════════════════════════════════════════
    (
        'TC-E-001', 'Edge Cases', 'High',
        '404 stránka pre neexistujúce URL',
        'Web dostupný',
        '1. Otvor https://www.mais.sk/sk/neexistujuca-stranka-xyz\n2. Over HTTP status (F12 → Network)\n3. Over obsah stránky\n4. Klikni na link späť na homepage',
        '404 stránka sa zobrazí. HTTP status 404. Obsahuje odkaz na homepage. Navbar funkčný.',
    ),
    (
        'TC-E-002', 'Edge Cases', 'Medium',
        'Prázdny formulár – validácia povinných polí',
        'Prihlásený do adminu, /admin/partners otvorená',
        '1. Klikni Pridať partnera\n2. Nič nevyplňuj\n3. Klikni Uložiť',
        'Validačné chyby sa zobrazia pri povinných poliach. Formulár sa neodošle. Žiadne prázdne záznamy v DB.',
    ),
    (
        'TC-E-003', 'Edge Cases', 'Low',
        'Veľmi dlhý text nepokazí layout',
        'Prihlásený do adminu',
        '1. Do CMS poľa zadaj text dlhší ako 500 znakov\n2. Ulož\n3. Otvor web na desktop (1366 px)\n4. Otvor web na mobile (375 px)\n5. Skontroluj overflow',
        'Layout sa nepokazí na žiadnom breakpointe. Žiadny horizontal overflow. Text zabalený alebo skrátený.',
    ),
    (
        'TC-E-004', 'Edge Cases', 'Medium',
        'Špeciálne a medzinárodné znaky v texte',
        'Prehliadač s podporou UTF-8',
        '1. Otvor /uk a over zobrazenie kyriliky\n2. Otvor /hu a over maďarské znaky (ő, ű, á)\n3. Over adresu mais@mais.sk\n4. Skontroluj DevTools Console – žiadne kódovacie chyby',
        'Všetky špeciálne znaky zobrazené správne. Žiadne "?" alebo rámčeky namiesto znakov.',
    ),
    (
        'TC-E-005', 'Edge Cases', 'Low',
        'Načítanie pri pomalom pripojení (Slow 3G)',
        'Chrome DevTools dostupné',
        '1. F12 → Network → Throttle: Slow 3G\n2. Načítaj hlavnú stránku (Ctrl+Shift+R)\n3. Sleduj loading state\n4. Over že stránka sa načíta\n5. Over DevTools Console',
        'Stránka sa načíta aj pri Slow 3G. Kľúčový obsah viditeľný po načítaní. Žiadne JS chyby.',
    ),

    # ══════════════════════════════════════════════════════════════════════════
    # PERFORMANCE (TC-P)
    # ══════════════════════════════════════════════════════════════════════════
    (
        'TC-P-001', 'Performance', 'High',
        'Lighthouse Desktop – skóre',
        'Chrome, produkčný web alebo beta (nie localhost)',
        '1. Otvor https://www.mais.sk/sk\n2. F12 → Lighthouse tab\n3. Vyber Desktop\n4. Klikni Analyze page load\n5. Zapíš všetky 4 skóre',
        'Performance ≥ 90. Accessibility ≥ 95. Best Practices ≥ 90. SEO ≥ 95.',
    ),
    (
        'TC-P-002', 'Performance', 'High',
        'Lighthouse Mobile – skóre',
        'Chrome, produkčný web alebo beta',
        '1. Otvor https://www.mais.sk/sk\n2. F12 → Lighthouse tab\n3. Vyber Mobile\n4. Klikni Analyze page load\n5. Zapíš skóre',
        'Performance ≥ 75. Accessibility ≥ 95. Best Practices ≥ 90. SEO ≥ 95.',
    ),
    (
        'TC-P-003', 'Performance', 'Medium',
        'First Contentful Paint (FCP) < 1.5 s',
        'Lighthouse Desktop spustené',
        '1. Spusti Lighthouse Desktop\n2. Nájdi metriku FCP (First Contentful Paint)',
        'FCP < 1.5 s na desktop. Prvý text alebo obrázok viditeľný rýchlo po načítaní.',
    ),
    (
        'TC-P-004', 'Performance', 'Medium',
        'Total Blocking Time (TBT) < 300 ms',
        'Lighthouse Desktop spustené',
        '1. Spusti Lighthouse Desktop\n2. Nájdi metriku TBT (Total Blocking Time)',
        'Total Blocking Time < 300 ms. Stránka je interaktívna rýchlo po načítaní.',
    ),
    (
        'TC-P-005', 'Performance', 'Low',
        'Bundle size JS < 500 KB compressed',
        'Chrome DevTools, produkčný build',
        '1. F12 → Network tab\n2. Načítaj stránku (Ctrl+Shift+R)\n3. Filter: JS\n4. Pozri stĺpec Transferred\n5. Spočítaj celkovú veľkosť',
        'Celková veľkosť JS súborov (Transferred = compressed) < 500 KB. Žiadny extrémne veľký bundle.',
    ),
]


# ─── Excel generátor ──────────────────────────────────────────────────────────

def _cell_style(cell, bg=WHITE, bold=False, font_color=MAIS_DARK, size=10):
    cell.font = Font(bold=bold, color=font_color, size=size, name='Calibri')
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    cell.border = BORDER


def create_excel(test_cases, output_path):
    wb = Workbook()

    # ── Sheet 1: Test Cases ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Test Cases'
    ws.sheet_view.showGridLines = False

    # Header row
    ws.row_dimensions[1].height = 40
    for col_idx, (header, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color=WHITE, size=11, name='Calibri')
        cell.fill = PatternFill('solid', fgColor=MAIS_ORANGE)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(COLUMNS))}1'

    # Dropdown validation for Stav column (col 9)
    dv = DataValidation(type='list', formula1='"Pass,Fail,Blocked,N/A"', showDropDown=False)
    dv.sqref = f'I2:I{len(test_cases) + 1}'
    ws.add_data_validation(dv)

    for row_idx, tc in enumerate(test_cases, 2):
        row_bg = LIGHT_GRAY if row_idx % 2 == 0 else WHITE
        tc_id, cat, priority, name, prereq, steps, expected = tc

        ws.row_dimensions[row_idx].height = 90

        for col_idx, val in enumerate(tc, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)

            if col_idx == 3:  # Priorita – špeciálna farba
                p_bg = PRIORITY_COLORS.get(val, row_bg)
                p_fc = PRIORITY_FONT_COLORS.get(val, MAIS_DARK)
                _cell_style(cell, bg=p_bg, bold=True, font_color=p_fc)
            elif col_idx == 1:  # ID – tučné
                _cell_style(cell, bg=row_bg, bold=True)
            elif col_idx == 4:  # Názov – tučné
                _cell_style(cell, bg=row_bg, bold=True)
            else:
                _cell_style(cell, bg=row_bg)

    # Conditional formatting – Stav stĺpec
    stav_range = f'I2:I{len(test_cases) + 1}'

    def _cond(formula_val, fg_color, font_color=MAIS_DARK, bold=False):
        fill = PatternFill(bgColor=fg_color)
        font = Font(color=font_color, bold=bold, name='Calibri')
        return FormulaRule(formula=[f'$I2="{formula_val}"'], fill=fill, font=font)

    ws.conditional_formatting.add(stav_range, _cond('Pass',    'C6EFCE', '276221', bold=True))
    ws.conditional_formatting.add(stav_range, _cond('Fail',    'FFC7CE', '9C0006', bold=True))
    ws.conditional_formatting.add(stav_range, _cond('Blocked', 'FFE5CC', '7A3800'))
    ws.conditional_formatting.add(stav_range, _cond('N/A',     'EDEDED', '666666'))

    # ── Sheet 2: Súhrn ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Súhrn')
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions['A'].width = 28
    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 22
    ws2.row_dimensions[1].height = 50

    # Title
    t = ws2.cell(row=1, column=1, value='maisweb – Testovací plán')
    t.font = Font(bold=True, size=20, color=MAIS_ORANGE, name='Calibri')
    t.alignment = Alignment(vertical='center')
    ws2.merge_cells('A1:C1')

    ws2.cell(row=2, column=1, value=f'Vygenerované: {datetime.now().strftime("%d.%m.%Y %H:%M")}').font = \
        Font(size=10, color='999999')

    def s2_section(row, text):
        c = ws2.cell(row=row, column=1, value=text)
        c.font = Font(bold=True, size=13, color=MAIS_ORANGE)
        return c

    def s2_row(row, label, formula_or_val):
        l = ws2.cell(row=row, column=1, value=label)
        l.font = Font(bold=True, size=11)
        v = ws2.cell(row=row, column=2, value=formula_or_val)
        v.font = Font(size=11)
        v.alignment = Alignment(horizontal='center')
        return v

    n = len(test_cases) + 1

    s2_section(4, 'PREHĽAD TESTOV')
    s2_row(5, 'Celkový počet:',   len(test_cases))
    s2_row(6, '🔴 Critical:',     f"=COUNTIF('Test Cases'!C:C,\"Critical\")")
    s2_row(7, '🟠 High:',         f"=COUNTIF('Test Cases'!C:C,\"High\")")
    s2_row(8, '🟡 Medium:',       f"=COUNTIF('Test Cases'!C:C,\"Medium\")")
    s2_row(9, '🟢 Low:',          f"=COUNTIF('Test Cases'!C:C,\"Low\")")

    s2_section(11, 'VÝSLEDKY TESTOVANIA')
    s2_row(12, '✅ Pass:',        f"=COUNTIF('Test Cases'!I2:I{n},\"Pass\")")
    s2_row(13, '❌ Fail:',        f"=COUNTIF('Test Cases'!I2:I{n},\"Fail\")")
    s2_row(14, '⚠️ Blocked:',    f"=COUNTIF('Test Cases'!I2:I{n},\"Blocked\")")
    s2_row(15, '➖ N/A:',         f"=COUNTIF('Test Cases'!I2:I{n},\"N/A\")")
    s2_row(16, '⬜ Bez výsledku:', f"=COUNTA('Test Cases'!A2:A{n})-COUNTA('Test Cases'!I2:I{n})")

    pct = ws2.cell(row=17, column=1, value='% splnenia (Pass / testované):')
    pct.font = Font(bold=True, size=11)
    pct_val = ws2.cell(row=17, column=2,
                        value=f"=IF(COUNTA('Test Cases'!I2:I{n})>0,"
                              f"COUNTIF('Test Cases'!I2:I{n},\"Pass\")/COUNTA('Test Cases'!I2:I{n}),0)")
    pct_val.number_format = '0%'
    pct_val.font = Font(bold=True, size=13, color='276221')

    # ── Sheet 3: Kategórie ───────────────────────────────────────────────────
    ws3 = wb.create_sheet('Kategórie')
    ws3.sheet_view.showGridLines = False
    ws3.row_dimensions[1].height = 36

    headers3 = ['Kategória', 'Celkom', 'Pass', 'Fail', 'Blocked', 'N/A', 'Bez výsledku', '% splnenia']
    widths3   = [18,          10,       10,     10,     12,        8,     16,             14]
    for col_idx, (h, w) in enumerate(zip(headers3, widths3), 1):
        cell = ws3.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True, color=WHITE, size=11, name='Calibri')
        cell.fill = PatternFill('solid', fgColor=MAIS_ORANGE)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER
        ws3.column_dimensions[get_column_letter(col_idx)].width = w

    categories = [
        'Funkčné', 'Vizuálne', 'Obsahové', 'Admin',
        'Content CMS', 'SEO', 'Security', 'Lokalizácia',
        'Edge Cases', 'Performance',
    ]

    for i, cat in enumerate(categories, 2):
        bg = LIGHT_GRAY if i % 2 == 0 else WHITE
        row_data = [
            cat,
            f"=COUNTIF('Test Cases'!B:B,\"{cat}\")",
            f"=COUNTIFS('Test Cases'!B:B,\"{cat}\",'Test Cases'!I:I,\"Pass\")",
            f"=COUNTIFS('Test Cases'!B:B,\"{cat}\",'Test Cases'!I:I,\"Fail\")",
            f"=COUNTIFS('Test Cases'!B:B,\"{cat}\",'Test Cases'!I:I,\"Blocked\")",
            f"=COUNTIFS('Test Cases'!B:B,\"{cat}\",'Test Cases'!I:I,\"N/A\")",
            f"=B{i}-C{i}-D{i}-E{i}-F{i}",
            f"=IF(B{i}>0,C{i}/B{i},0)",
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws3.cell(row=i, column=col_idx, value=val)
            cell.fill = PatternFill('solid', fgColor=bg)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal='center' if col_idx > 1 else 'left',
                                        vertical='center')
            cell.font = Font(size=10, name='Calibri', bold=(col_idx == 1))
            if col_idx == 8:
                cell.number_format = '0%'

    ws3.row_dimensions[1].height = 36

    wb.save(output_path)
    print(f'  ✓ Excel: {output_path}')


# ─── Markdown generátor ──────────────────────────────────────────────────────

CATEGORY_ORDER = [
    ('Funkčné',     '1. Funkčné testy'),
    ('Vizuálne',    '2. Vizuálne testy'),
    ('Obsahové',    '3. Obsahové testy'),
    ('Admin',       '4. Admin testy'),
    ('Content CMS', '5. Content CMS testy'),
    ('SEO',         '6. SEO testy'),
    ('Security',    '7. Security testy'),
    ('Lokalizácia', '8. Lokalizácia testy'),
    ('Edge Cases',  '9. Edge Cases'),
    ('Performance', '10. Performance testy'),
]

PRIORITY_EMOJI = {
    'Critical': '🔴',
    'High':     '🟠',
    'Medium':   '🟡',
    'Low':      '🟢',
}


def create_markdown(test_cases, output_path):
    by_cat = {}
    for tc in test_cases:
        by_cat.setdefault(tc[1], []).append(tc)

    lines = [
        '# Test cases pre maisweb',
        '',
        '> Verzia 1.0 · 27. apríla 2026 · ITernal s.r.o.',
        '',
        '## Ako používať',
        '',
        '- Každý krok testu je zaškrtávateľný checkbox — zaškrtni po overení',
        '- Pri zlyhaní zaznamenaj do `test-cases.xlsx` (stĺpec Skutočný výsledok, Stav = Fail)',
        '- Vytvor GitHub Issue s ID testu (napr. `TC-F-003`) a prilož screenshot',
        '- **Priorita**: 🔴 Critical > 🟠 High > 🟡 Medium > 🟢 Low',
        '',
        '---',
        '',
    ]

    # Obsah
    lines.append('## Obsah')
    lines.append('')
    for cat_key, section_title in CATEGORY_ORDER:
        if cat_key not in by_cat:
            continue
        anchor = section_title.lower().replace(' ', '-').replace('.', '').replace('/', '')
        count = len(by_cat[cat_key])
        lines.append(f'- [{section_title}](#{anchor}) — {count} testov')
    lines.append('')
    lines.append('---')
    lines.append('')

    for cat_key, section_title in CATEGORY_ORDER:
        if cat_key not in by_cat:
            continue

        lines.append(f'## {section_title}')
        lines.append('')

        for tc in by_cat[cat_key]:
            tc_id, _, priority, name, prereq, steps, expected = tc
            emoji = PRIORITY_EMOJI.get(priority, '')

            lines.append(f'### {tc_id}: {name}')
            lines.append(f'**Priorita:** {emoji} {priority}')
            lines.append('')

            if prereq:
                lines.append(f'**Predpoklad:** {prereq}')
                lines.append('')

            lines.append('**Postup:**')
            for step in steps.split('\n'):
                lines.append(f'- [ ] {step}')
            lines.append('')

            lines.append(f'**Očakávaný výsledok:** {expected}')
            lines.append('')
            lines.append('**Skutočný výsledok:** _(vyplní tester)_')
            lines.append('')
            lines.append('**Stav:** `Pass` / `Fail` / `Blocked` / `N/A`')
            lines.append('')
            lines.append('---')
            lines.append('')

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))
    print(f'  ✓ Markdown: {output_path}')


# ─── Main ─────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    os.makedirs('docs/testing', exist_ok=True)

    xlsx_path = 'docs/testing/test-cases.xlsx'
    md_path   = 'docs/testing/test-cases.md'

    print(f'\nGenerujem {len(TEST_CASES)} test cases...')
    create_excel(TEST_CASES, xlsx_path)
    create_markdown(TEST_CASES, md_path)
    print(f'  Hotovo! ({len(TEST_CASES)} testov)\n')
